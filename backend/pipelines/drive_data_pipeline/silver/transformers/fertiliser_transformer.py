"""
Fertiliser transformer for processing fertiliser data files in the drive pipeline.

This transformer handles the harmonization of Danish fertiliser data from multiple sources:
- Efterafgrøder (cover crops)
- GKEA markplan files
- Gødningsregnskaber (fertilizer accounts)

Refactored to use vanilla DuckDB instead of pandas.
"""

import contextlib
import os
import tempfile
from pathlib import Path
from typing import Any

import duckdb
from common.logging_utils import get_pipeline_logger

from ..models.schema import ColumnSchema, DataType, TableSchema
from .base import BaseTransformer, TransformResult

# Handle imports for both standalone and package usage
try:
    from common.storage import get_duckdb_with_r2
except ImportError:
    # Fallback for standalone usage
    def get_duckdb_with_r2() -> duckdb.DuckDBPyConnection:
        conn = duckdb.connect()
        try:
            conn.execute("INSTALL spatial")
            conn.execute("LOAD spatial")
        except Exception:
            pass
        return conn


logger = get_pipeline_logger(__name__)


class FertiliserTransformer(BaseTransformer):
    """Transformer for fertiliser parquet files using vanilla DuckDB."""

    def __init__(self) -> None:
        """Initialize the fertiliser transformer."""
        super().__init__()
        self.conn = get_duckdb_with_r2()
        self._setup_harmonization_schemas()

    def _setup_harmonization_schemas(self) -> None:
        """Setup the standardized schemas for harmonized fertiliser data."""
        self.harmonized_schema = {
            "data_source": DataType.STRING,
            "year": DataType.STRING,
            "cvr_number": DataType.STRING,
            "capnumber": DataType.STRING,
            "markbloknummer": DataType.STRING,
            "marknummer": DataType.STRING,
            "indberet_alternativ": DataType.STRING,
            "faktisk_areal_ha": DataType.FLOAT,
            "omregnet_areal_ha": DataType.FLOAT,
            "journal_nummer": DataType.STRING,
            "total_n_kvote": DataType.FLOAT,
            "fosfortal": DataType.FLOAT,
            "data_type": DataType.STRING,
            "data_source_file": DataType.STRING,
        }

    def can_handle(self, file_path: Path, metadata: dict[str, Any]) -> bool:
        """
        Check if this transformer can handle the given file.

        Args:
            file_path: Path to the file
            metadata: File metadata

        Returns:
            True if this transformer can handle the file
        """
        filename = file_path.name.lower()

        # Get the full path to check for In-depth directory structure
        full_path_str = str(file_path).lower()

        # Check for main fertiliser-related files
        fertiliser_patterns = [
            "efterafgrøder",
            "efterafgroeder",
            "gkea",
            "gødningsregnskaber",
            "goedningsregnskaber",
            "fertiliser",
            "fertilizer",
        ]

        # Check for In-depth fertilizer files (B_*, V_*, etc. in GR folders)
        indepth_patterns = [
            "b_aftrk",
            "b_aoggoed",
            "b_biomasr",
            "b_blandrk",
            "b_dyrerk",
            "b_forarbr",
            "b_goedrk",
            "b_modhumr",
            "b_modrk",
            "b_ovdrk",
            "v_",  # V_ files
            "erklrk",
            "lg_company",
        ]

        # Check if it's in an In-depth/GR folder structure
        is_in_gr_folder = "in-depth" in full_path_str and (
            "gr " in full_path_str or "gr/" in full_path_str
        )

        # Main patterns match
        main_match = any(pattern in filename for pattern in fertiliser_patterns)

        # In-depth patterns match (and in correct folder structure)
        indepth_match = is_in_gr_folder and any(pattern in filename for pattern in indepth_patterns)

        return main_match or indepth_match

    def transform(
        self,
        file_path: Path,
        metadata: Any,
        output_dir: Path,
    ) -> TransformResult:
        """
        Transform a fertiliser file.

        Args:
            file_path: Path to the input file
            metadata: File metadata
            output_dir: Output directory for transformed data

        Returns:
            TransformResult containing the transformation outcome
        """
        try:
            logger.info(f"Transforming fertiliser file: {file_path.name}")

            # Read the file based on its extension
            file_suffix = file_path.suffix.lower()

            if file_suffix in (".xlsx", ".xls"):
                # Read Excel file using DuckDB with spatial extension for xlsx support
                table_name = self._read_excel_to_table(file_path)
                if table_name is None:
                    return TransformResult(success=False, error="Failed to read Excel file")
            elif file_suffix == ".parquet":
                # Read parquet file directly into DuckDB
                table_name = f"fertiliser_raw_{id(self)}"
                self.conn.execute(f"""
                    CREATE TABLE {table_name} AS
                    SELECT * FROM read_parquet('{file_path}')
                """)
            else:
                return TransformResult(
                    success=False,
                    error=f"Unsupported file type for fertiliser transformer: {file_suffix}",
                )

            # Check if table is empty
            row_count = self.conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            if row_count == 0:
                self._drop_table(table_name)
                return TransformResult(success=False, error="Empty data file")

            # Determine the type of fertiliser file and harmonize
            filename = file_path.name
            harmonized_table = self._harmonize_fertiliser_data(table_name, filename)

            if harmonized_table is None:
                self._drop_table(table_name)
                return TransformResult(success=False, error="No valid data after harmonization")

            # Get row count from harmonized table
            harmonized_count = self.conn.execute(
                f"SELECT COUNT(*) FROM {harmonized_table}"
            ).fetchone()[0]

            if harmonized_count == 0:
                self._drop_table(table_name)
                self._drop_table(harmonized_table)
                return TransformResult(success=False, error="No valid data after harmonization")

            # Create output path
            output_path = output_dir / f"{file_path.stem}_harmonized.parquet"

            # Ensure output directory exists
            output_dir.mkdir(parents=True, exist_ok=True)

            # Save harmonized data using DuckDB COPY
            self.conn.execute(f"""
                COPY {harmonized_table} TO '{output_path}' (FORMAT PARQUET)
            """)

            # Get column names for metadata
            columns_info = self.conn.execute(f"DESCRIBE {harmonized_table}").fetchall()
            harmonized_columns = [col[0] for col in columns_info]

            logger.info(
                f"Successfully transformed fertiliser data to: {output_path} "
                f"({harmonized_count} rows)"
            )

            # Cleanup tables
            self._drop_table(table_name)
            self._drop_table(harmonized_table)

            return TransformResult(
                success=True,
                output_path=output_path,
                row_count=harmonized_count,
                metadata={
                    "original_filename": filename,
                    "harmonized_columns": harmonized_columns,
                    "data_source": self._get_data_source(filename),
                    "data_type": self._get_data_type(filename),
                },
            )

        except Exception as e:
            logger.error(f"Failed to transform fertiliser file {file_path}: {e!s}")
            return TransformResult(success=False, error=str(e))

    def _read_excel_to_table(self, file_path: Path) -> str | None:
        """Read Excel file into a DuckDB table.

        Uses DuckDB's spatial extension for xlsx reading, with fallback to
        temporary CSV conversion.

        Args:
            file_path: Path to the Excel file

        Returns:
            Table name or None if failed
        """
        table_name = f"excel_raw_{id(self)}"

        try:
            # Try using DuckDB's spatial extension st_read for xlsx
            # This requires the spatial extension
            self.conn.execute("INSTALL spatial; LOAD spatial;")
            self.conn.execute(f"""
                CREATE TABLE {table_name} AS
                SELECT * FROM st_read('{file_path}')
            """)
            logger.info(f"Read Excel file using DuckDB spatial extension: {file_path.name}")
            return table_name
        except Exception as spatial_e:
            logger.debug(f"Spatial extension xlsx read failed: {spatial_e}")

        # Fallback: Use Python's openpyxl to read and convert to CSV, then load
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            all_data = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                # First row as headers
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

                for row in rows[1:]:
                    row_dict = {
                        headers[i]: str(v) if v is not None else "" for i, v in enumerate(row)
                    }
                    row_dict["source_sheet"] = sheet_name
                    all_data.append(row_dict)

            wb.close()

            if not all_data:
                logger.warning(f"No data found in Excel file: {file_path.name}")
                return None

            # Get all unique columns
            all_columns = set()
            for row in all_data:
                all_columns.update(row.keys())
            all_columns = sorted(all_columns)

            # Create table with all columns as VARCHAR
            columns_def = ", ".join([f'"{col}" VARCHAR' for col in all_columns])
            self.conn.execute(f"CREATE TABLE {table_name} ({columns_def})")

            # Insert data row by row
            for row in all_data:
                values = []
                for col in all_columns:
                    val = row.get(col, "")
                    # Escape single quotes
                    val = str(val).replace("'", "''")
                    values.append(f"'{val}'")
                values_str = ", ".join(values)
                self.conn.execute(f"INSERT INTO {table_name} VALUES ({values_str})")

            logger.info(
                f"Read Excel file using openpyxl fallback: {file_path.name} ({len(all_data)} rows)"
            )
            return table_name

        except ImportError:
            logger.error("openpyxl not installed - cannot read Excel files")
            return None
        except Exception as e:
            logger.error(f"Failed to read Excel file {file_path.name}: {e!s}")
            return None

    def transform_from_content(
        self, content: bytes, filename: str, metadata: dict[str, Any]
    ) -> Any:
        """
        Transform fertiliser data from file content.

        Args:
            content: File content bytes
            filename: Original filename
            metadata: File metadata

        Returns:
            DataFrame with harmonized data or None if transformation failed
        """
        try:
            # Get the original file extension
            file_suffix = Path(filename).suffix.lower()

            # Create temporary file with correct extension
            with tempfile.NamedTemporaryFile(suffix=file_suffix, delete=False) as tmp:
                tmp.write(content)
                tmp.flush()
                tmp_path = tmp.name

            try:
                # Read the file based on its extension
                if file_suffix in (".xlsx", ".xls"):
                    table_name = self._read_excel_to_table(Path(tmp_path))
                    if table_name is None:
                        logger.warning(f"No readable sheets found in Excel file: {filename}")
                        return None
                elif file_suffix == ".parquet":
                    table_name = f"fertiliser_content_{id(self)}"
                    self.conn.execute(f"""
                        CREATE TABLE {table_name} AS
                        SELECT * FROM read_parquet('{tmp_path}')
                    """)
                else:
                    logger.error(f"Unsupported file type for fertiliser transformer: {file_suffix}")
                    return None

                # Check if empty
                row_count = self.conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                if row_count == 0:
                    logger.warning(f"Empty data file: {filename}")
                    self._drop_table(table_name)
                    return None

                # Harmonize the data
                harmonized_table = self._harmonize_fertiliser_data(table_name, filename)

                if harmonized_table is None:
                    logger.warning(f"No data after harmonization for: {filename}")
                    self._drop_table(table_name)
                    return None

                harmonized_count = self.conn.execute(
                    f"SELECT COUNT(*) FROM {harmonized_table}"
                ).fetchone()[0]

                if harmonized_count > 0:
                    logger.info(
                        f"Successfully harmonized fertiliser data from: {filename} "
                        f"({harmonized_count} rows)"
                    )
                    # Return as DataFrame for compatibility with existing code
                    result_df = self.conn.execute(f"SELECT * FROM {harmonized_table}").df()
                    self._drop_table(table_name)
                    self._drop_table(harmonized_table)
                    return result_df

                logger.warning(f"No data after harmonization for: {filename}")
                self._drop_table(table_name)
                self._drop_table(harmonized_table)
                return None

            finally:
                # Clean up temporary file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        except Exception as e:
            logger.error(f"Failed to transform fertiliser content from {filename}: {e!s}")
            return None

    def _drop_table(self, table_name: str) -> None:
        """Safely drop a table if it exists."""
        with contextlib.suppress(Exception):
            self.conn.execute(f"DROP TABLE IF EXISTS {table_name}")

    def _harmonize_fertiliser_data(self, table_name: str, filename: str) -> str | None:
        """
        Harmonize fertiliser data based on the file type.

        Args:
            table_name: Name of the DuckDB table with raw data
            filename: Source filename to determine processing type

        Returns:
            Name of harmonized table or None if failed
        """
        try:
            filename_lower = filename.lower()

            if "efterafgrøder" in filename_lower or "efterafgroeder" in filename_lower:
                return self._process_efterafgroeder(table_name, filename)
            if "gkea" in filename_lower:
                return self._process_gkea(table_name, filename)
            if (
                "gødningsregnskaber" in filename_lower
                or "goedningsregnskaber" in filename_lower
                or any(
                    pattern in filename_lower for pattern in ["b_", "v_", "erklrk", "lg_company"]
                )
            ):
                return self._process_goedningsregnskaber(table_name, filename)
            logger.warning(f"Unknown fertiliser file type: {filename}")
            return self._process_generic_fertiliser(table_name, filename)

        except Exception as e:
            logger.error(f"Failed to harmonize fertiliser data from {filename}: {e!s}")
            return None

    def _get_column_if_exists(self, table_name: str, column_name: str) -> str:
        """Return column reference or NULL if column doesn't exist."""
        columns = self.conn.execute(f"DESCRIBE {table_name}").fetchall()
        column_names = [col[0] for col in columns]
        if column_name in column_names:
            return f'"{column_name}"'
        return "NULL"

    def _process_efterafgroeder(self, table_name: str, filename: str) -> str:
        """Process Efterafgrøder (cover crops) files using DuckDB."""
        logger.info(f"Processing Efterafgrøder file: {filename}")

        harmonized_table = f"harmonized_efterafgroeder_{id(self)}"

        # Get available columns
        columns = self.conn.execute(f"DESCRIBE {table_name}").fetchall()
        column_names = [col[0] for col in columns]

        # Build dynamic column selections based on what's available
        year_col = self._get_column_if_exists(table_name, "PROD_AAR")
        cvr_col = self._get_column_if_exists(table_name, "CVR")
        cap_col = self._get_column_if_exists(table_name, "CapNumber")
        markblok_col = self._get_column_if_exists(table_name, "MARKBLOKNUMMER")
        marknr_col = self._get_column_if_exists(table_name, "MARKNUMMER")

        # Determine which year format we have for alternativ columns
        indberet_col = "NULL"
        faktisk_col = "NULL"
        omregnet_col = "NULL"

        # Check for different year formats
        if "A19_INDBERETEFTERAFGALTERNATIV" in column_names:
            # 2023 format
            indberet_col = '"A19_INDBERETEFTERAFGALTERNATIV"'
            if "A20_FAKTISKHAUDLAGTEAALTERNATIV" in column_names:
                faktisk_col = """TRY_CAST(
                    REPLACE(CAST("A20_FAKTISKHAUDLAGTEAALTERNATIV" AS VARCHAR), ',', '.')
                    AS DOUBLE)"""
            if "A23_OMREGNETHAMEDEA" in column_names:
                omregnet_col = """TRY_CAST(
                    REPLACE(CAST("A23_OMREGNETHAMEDEA" AS VARCHAR), ',', '.')
                    AS DOUBLE)"""
        elif "A18_INDBERETEFTERAFGALTERNATIV" in column_names:
            # 2020 format
            indberet_col = '"A18_INDBERETEFTERAFGALTERNATIV"'
            if "A19_FAKTISKHAUDLAGTEAALTERNATIV" in column_names:
                faktisk_col = """TRY_CAST(
                    REPLACE(CAST("A19_FAKTISKHAUDLAGTEAALTERNATIV" AS VARCHAR), ',', '.')
                    AS DOUBLE)"""
            if "A20_OMREGNETHAMEDEA" in column_names:
                omregnet_col = """TRY_CAST(
                    REPLACE(CAST("A20_OMREGNETHAMEDEA" AS VARCHAR), ',', '.')
                    AS DOUBLE)"""
        elif "A20_INDBERETEFTERAFGALTERNATIV" in column_names:
            # 2022 format
            indberet_col = '"A20_INDBERETEFTERAFGALTERNATIV"'
            if "A21_FAKTISKHAUDLAGTEAALTERNATIV" in column_names:
                faktisk_col = """TRY_CAST(
                    REPLACE(CAST("A21_FAKTISKHAUDLAGTEAALTERNATIV" AS VARCHAR), ',', '.')
                    AS DOUBLE)"""
            if "A24_OMREGNETHAMEDEA" in column_names:
                omregnet_col = """TRY_CAST(
                    REPLACE(CAST("A24_OMREGNETHAMEDEA" AS VARCHAR), ',', '.')
                    AS DOUBLE)"""

        # Escape filename for SQL
        escaped_filename = filename.replace("'", "''")

        self.conn.execute(f"""
            CREATE TABLE {harmonized_table} AS
            SELECT
                'efterafgroeder' AS data_source,
                CAST({year_col} AS VARCHAR) AS year,
                CAST({cvr_col} AS VARCHAR) AS cvr_number,
                CAST({cap_col} AS VARCHAR) AS capnumber,
                CAST({markblok_col} AS VARCHAR) AS markbloknummer,
                CAST({marknr_col} AS VARCHAR) AS marknummer,
                CAST({indberet_col} AS VARCHAR) AS indberet_alternativ,
                {faktisk_col} AS faktisk_areal_ha,
                {omregnet_col} AS omregnet_areal_ha,
                NULL AS journal_nummer,
                NULL::DOUBLE AS total_n_kvote,
                NULL::DOUBLE AS fosfortal,
                'Efterafgrøder' AS data_type,
                '{escaped_filename}' AS data_source_file
            FROM {table_name}
        """)

        return harmonized_table

    def _process_gkea(self, table_name: str, filename: str) -> str:
        """Process GKEA markplan files using DuckDB."""
        logger.info(f"Processing GKEA file: {filename}")

        harmonized_table = f"harmonized_gkea_{id(self)}"

        # Extract year from filename
        year = None
        for yr in ["2021", "2022", "2023", "2024"]:
            if yr in filename:
                year = yr
                break

        # Get column info
        columns = self.conn.execute(f"DESCRIBE {table_name}").fetchall()
        num_cols = len(columns)

        # Escape filename for SQL
        escaped_filename = filename.replace("'", "''")

        # GKEA files typically have positional columns
        # Create the harmonized table based on column positions
        if num_cols >= 2:
            # Build column references by position (0-indexed)
            # First, rename columns to positional names for easier access
            select_parts = []
            for i, col in enumerate(columns):
                col_name = col[0]
                escaped_name = col_name.replace('"', '""')
                select_parts.append(f'"{escaped_name}" AS column{i}')

            positional_table = f"gkea_positional_{id(self)}"
            self.conn.execute(f"""
                CREATE TABLE {positional_table} AS
                SELECT {", ".join(select_parts)}
                FROM {table_name}
            """)

            # Now create harmonized table based on year-specific mappings
            journal_col = "column0" if num_cols > 0 else "NULL"
            cvr_col = "column1" if num_cols > 1 else "NULL"

            if year == "2021" and num_cols > 14:
                marknr_col = "column5"
                faktisk_col = "TRY_CAST(REPLACE(CAST(column6 AS VARCHAR), ',', '.') AS DOUBLE)"
                omregnet_col = "TRY_CAST(REPLACE(CAST(column10 AS VARCHAR), ',', '.') AS DOUBLE)"
                indberet_col = "column14"
                fosfor_col = (
                    "TRY_CAST(REPLACE(CAST(column19 AS VARCHAR), ',', '.') AS DOUBLE)"
                    if num_cols > 19
                    else "NULL"
                )
            elif year == "2022" and num_cols > 12:
                marknr_col = "column3"
                faktisk_col = "TRY_CAST(REPLACE(CAST(column4 AS VARCHAR), ',', '.') AS DOUBLE)"
                omregnet_col = "TRY_CAST(REPLACE(CAST(column8 AS VARCHAR), ',', '.') AS DOUBLE)"
                indberet_col = "column12"
                fosfor_col = (
                    "TRY_CAST(REPLACE(CAST(column17 AS VARCHAR), ',', '.') AS DOUBLE)"
                    if num_cols > 17
                    else "NULL"
                )
            elif year in ["2023", "2024"] and num_cols > 10:
                marknr_col = "column3"
                faktisk_col = "TRY_CAST(REPLACE(CAST(column4 AS VARCHAR), ',', '.') AS DOUBLE)"
                omregnet_col = "TRY_CAST(REPLACE(CAST(column6 AS VARCHAR), ',', '.') AS DOUBLE)"
                indberet_col = "column10"
                fosfor_col = "NULL"
            else:
                marknr_col = "NULL"
                faktisk_col = "NULL"
                omregnet_col = "NULL"
                indberet_col = "NULL"
                fosfor_col = "NULL"

            year_val = f"'{year}'" if year else "NULL"

            self.conn.execute(f"""
                CREATE TABLE {harmonized_table} AS
                SELECT
                    'gkea' AS data_source,
                    {year_val} AS year,
                    CAST({cvr_col} AS VARCHAR) AS cvr_number,
                    NULL AS capnumber,
                    NULL AS markbloknummer,
                    CAST({marknr_col} AS VARCHAR) AS marknummer,
                    CAST({indberet_col} AS VARCHAR) AS indberet_alternativ,
                    {faktisk_col} AS faktisk_areal_ha,
                    {omregnet_col} AS omregnet_areal_ha,
                    CAST({journal_col} AS VARCHAR) AS journal_nummer,
                    NULL::DOUBLE AS total_n_kvote,
                    {fosfor_col} AS fosfortal,
                    'GKEA Markplan' AS data_type,
                    '{escaped_filename}' AS data_source_file
                FROM {positional_table}
            """)

            # Clean up positional table
            self._drop_table(positional_table)
        else:
            # Not enough columns, create empty harmonized table
            self.conn.execute(f"""
                CREATE TABLE {harmonized_table} AS
                SELECT
                    'gkea' AS data_source,
                    NULL AS year,
                    NULL AS cvr_number,
                    NULL AS capnumber,
                    NULL AS markbloknummer,
                    NULL AS marknummer,
                    NULL AS indberet_alternativ,
                    NULL::DOUBLE AS faktisk_areal_ha,
                    NULL::DOUBLE AS omregnet_areal_ha,
                    NULL AS journal_nummer,
                    NULL::DOUBLE AS total_n_kvote,
                    NULL::DOUBLE AS fosfortal,
                    'GKEA Markplan' AS data_type,
                    '{escaped_filename}' AS data_source_file
                WHERE 1=0
            """)

        return harmonized_table

    def _process_goedningsregnskaber(self, table_name: str, filename: str) -> str:
        """Process Gødningsregnskaber (fertilizer accounts) files using DuckDB."""
        logger.info(f"Processing Gødningsregnskaber/In-depth file: {filename}")

        harmonized_table = f"harmonized_goedning_{id(self)}"

        # Extract year from filename
        year = None
        for yr in ["2018", "2019", "2020", "2021", "2022", "2023", "2024"]:
            if yr in filename:
                year = yr
                break

        # Determine data source and type
        filename_lower = filename.lower()
        if "gødningsregnskaber" in filename_lower or "goedningsregnskaber" in filename_lower:
            data_source = "goedningsregnskaber"
            data_type = "Gødningsregnskaber"
        elif "b_" in filename_lower:
            data_source = "goedningsregnskaber_blok"
            data_type = "Gødningsregnskaber Blok"
        elif "v_" in filename_lower:
            data_source = "goedningsregnskaber_vurdering"
            data_type = "Gødningsregnskaber Vurdering"
        else:
            data_source = "goedningsregnskaber_other"
            data_type = "Gødningsregnskaber Other"

        # Get column mapping
        cvr_col = self._get_column_if_exists(table_name, "CVR")
        nummer_col = self._get_column_if_exists(table_name, "NUMMER")

        year_val = f"'{year}'" if year else "NULL"

        # Escape filename for SQL
        escaped_filename = filename.replace("'", "''")

        self.conn.execute(f"""
            CREATE TABLE {harmonized_table} AS
            SELECT
                '{data_source}' AS data_source,
                {year_val} AS year,
                CAST({cvr_col} AS VARCHAR) AS cvr_number,
                NULL AS capnumber,
                NULL AS markbloknummer,
                NULL AS marknummer,
                NULL AS indberet_alternativ,
                NULL::DOUBLE AS faktisk_areal_ha,
                NULL::DOUBLE AS omregnet_areal_ha,
                CAST({nummer_col} AS VARCHAR) AS journal_nummer,
                NULL::DOUBLE AS total_n_kvote,
                NULL::DOUBLE AS fosfortal,
                '{data_type}' AS data_type,
                '{escaped_filename}' AS data_source_file
            FROM {table_name}
        """)

        return harmonized_table

    def _process_generic_fertiliser(self, table_name: str, filename: str) -> str:
        """Process generic fertiliser files using DuckDB."""
        logger.info(f"Processing generic fertiliser file: {filename}")

        harmonized_table = f"harmonized_generic_{id(self)}"

        # Check for cvr_number column
        cvr_col = self._get_column_if_exists(table_name, "cvr_number")

        # Escape filename for SQL
        escaped_filename = filename.replace("'", "''")

        self.conn.execute(f"""
            CREATE TABLE {harmonized_table} AS
            SELECT
                'generic' AS data_source,
                NULL AS year,
                CAST({cvr_col} AS VARCHAR) AS cvr_number,
                NULL AS capnumber,
                NULL AS markbloknummer,
                NULL AS marknummer,
                NULL AS indberet_alternativ,
                NULL::DOUBLE AS faktisk_areal_ha,
                NULL::DOUBLE AS omregnet_areal_ha,
                NULL AS journal_nummer,
                NULL::DOUBLE AS total_n_kvote,
                NULL::DOUBLE AS fosfortal,
                'Generic Fertiliser' AS data_type,
                '{escaped_filename}' AS data_source_file
            FROM {table_name}
        """)

        return harmonized_table

    def _get_data_source(self, filename: str) -> str:
        """Determine data source from filename."""
        filename_lower = filename.lower()

        if "efterafgrøder" in filename_lower or "efterafgroeder" in filename_lower:
            return "efterafgroeder"
        if "gkea" in filename_lower:
            return "gkea"
        if "gødningsregnskaber" in filename_lower or "goedningsregnskaber" in filename_lower:
            return "goedningsregnskaber"
        if "b_" in filename_lower:
            return "goedningsregnskaber_blok"
        if "v_" in filename_lower:
            return "goedningsregnskaber_vurdering"
        return "generic"

    def _get_data_type(self, filename: str) -> str:
        """Determine data type from filename."""
        filename_lower = filename.lower()

        if "efterafgrøder" in filename_lower or "efterafgroeder" in filename_lower:
            return "Efterafgrøder"
        if "gkea" in filename_lower:
            return "GKEA Markplan"
        if "gødningsregnskaber" in filename_lower or "goedningsregnskaber" in filename_lower:
            return "Gødningsregnskaber"
        if "b_" in filename_lower:
            return "Gødningsregnskaber Blok"
        if "v_" in filename_lower:
            return "Gødningsregnskaber Vurdering"
        return "Generic Fertiliser"

    def get_expected_schema(self) -> TableSchema | None:
        """
        Get the expected schema for harmonized fertiliser data.

        Returns:
            TableSchema for the harmonized output
        """
        columns = []
        for col_name, col_type in self.harmonized_schema.items():
            columns.append(
                ColumnSchema(
                    name=col_name,
                    data_type=col_type,
                    nullable=True,  # Most columns are nullable in fertiliser data
                    description=f"Harmonized {col_name} column",
                )
            )

        return TableSchema(
            name="fertiliser_harmonized",
            columns=columns,
            description="Harmonized Danish fertiliser data from multiple sources",
        )
