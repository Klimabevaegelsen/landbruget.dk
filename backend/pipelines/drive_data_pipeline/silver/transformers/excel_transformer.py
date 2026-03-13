"""Excel file transformer using DuckDB for processing.

Refactored to use vanilla DuckDB instead of pandas, with fallback to openpyxl
for reading Excel files when DuckDB's spatial extension is unavailable.
"""

import time
from pathlib import Path

# Handle imports for both standalone and package usage
try:
    from ...utils.logging import get_logger
    from ..duckdb_base import DuckDBProcessor
    from .base import BaseTransformer, FileMetadata, TransformResult
except ImportError:
    # Fallback for standalone usage
    import logging

    def get_logger() -> logging.Logger:
        return logging.getLogger(__name__)

    from silver.duckdb_base import DuckDBProcessor
    from silver.transformers.base import BaseTransformer, FileMetadata, TransformResult

logger = get_logger()


class ExcelTransformer(BaseTransformer, DuckDBProcessor):
    """Transform Excel files using DuckDB for processing."""

    def __init__(self) -> None:
        BaseTransformer.__init__(self)
        DuckDBProcessor.__init__(self)
        logger.info("Initialized ExcelTransformer with DuckDB")

    def transform(
        self,
        file_path: Path,
        metadata: FileMetadata,
        output_dir: Path,
    ) -> TransformResult:
        """Transform Excel file to standardized format using DuckDB.

        Args:
            file_path: Path to the Excel file
            metadata: File metadata
            output_dir: Output directory for transformed files

        Returns:
            TransformResult with success status and metadata
        """
        try:
            logger.info(f"Transforming Excel file using DuckDB: {file_path}")

            # Create output directory for this file
            file_output_dir = output_dir / "Excel"
            file_output_dir.mkdir(parents=True, exist_ok=True)

            # Read Excel file and get all sheets
            sheets_data = self._read_excel_sheets(file_path)

            if not sheets_data:
                logger.warning(f"No valid sheets found in Excel file: {file_path}")
                return TransformResult(
                    success=False,
                    error="No valid sheets found in Excel file",
                )

            output_paths = []
            total_rows = 0
            last_standardized_table = None
            standardized_tables = []  # Track all standardized tables for cleanup

            for sheet_name, table_name in sheets_data:
                # Clean column names using DuckDB
                cleaned_table = self._standardize_column_names_duckdb(table_name)

                # Apply data type standardization using DuckDB
                standardized_table = self._standardize_data_types_duckdb(cleaned_table)

                # Keep track of standardized tables
                last_standardized_table = standardized_table
                standardized_tables.append(standardized_table)

                # Generate output filename
                base_filename = Path(metadata.original_filename).stem
                sheet_filename = f"{base_filename}_{sheet_name}"

                # Save as Parquet using DuckDB
                output_path = file_output_dir / f"{sheet_filename}.parquet"
                self.save_table_to_parquet(standardized_table, output_path)

                output_paths.append(output_path)

                # Get row count
                row_count = self.conn.execute(
                    f"SELECT COUNT(*) FROM {standardized_table}"
                ).fetchone()[0]
                total_rows += row_count

                # Clean up intermediate tables (but keep standardized tables for schema)
                self.drop_table(table_name)
                # Only drop cleaned_table if it's different from standardized_table
                if cleaned_table != standardized_table:
                    self.drop_table(cleaned_table)

            # Create schema dictionary from the last table (for compatibility)
            schema = {"columns": [], "data_types": []}
            if sheets_data and last_standardized_table:
                # Check if the table still exists before trying to get schema
                try:
                    table_info = self.get_table_info(last_standardized_table)
                    schema = {
                        "columns": [col[0] for col in table_info],
                        "data_types": [col[1] for col in table_info],
                    }
                except Exception as e:
                    logger.warning(
                        f"Failed to get schema from table {last_standardized_table}: {e!s}"
                    )
                    # Use basic schema as fallback
                    schema = {"columns": [], "data_types": []}

                # Now clean up all standardized tables
                for standardized_table in standardized_tables:
                    self.drop_table(standardized_table)

            return TransformResult(
                success=True,
                output_path=output_paths[0] if len(output_paths) == 1 else None,
                row_count=total_rows,
                schema=schema,
                metadata={
                    "sheet_count": len(sheets_data),
                    "output_paths": [str(p) for p in output_paths],
                },
            )

        except Exception as e:
            error_msg = f"Failed to transform Excel file {file_path}: {e!s}"
            logger.error(error_msg)
            return TransformResult(
                success=False,
                error=error_msg,
            )

    def _read_excel_sheets(self, file_path: Path) -> list[tuple[str, str]]:
        """Read Excel file sheets using DuckDB with fallback to openpyxl.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of tuples containing (sheet_name, table_name)
        """
        # Try using DuckDB's spatial extension first for xlsx support
        try:
            self.conn.execute("INSTALL spatial; LOAD spatial;")
            return self._read_excel_with_spatial(file_path)
        except Exception as e:
            logger.debug(f"Spatial extension not available or failed: {e}")

        # Fallback to openpyxl
        return self._read_excel_with_openpyxl(file_path)

    def _read_excel_with_spatial(self, file_path: Path) -> list[tuple[str, str]]:
        """Read Excel file using DuckDB's spatial extension st_read.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of tuples containing (sheet_name, table_name)
        """
        sheets_data = []

        # st_read with layer parameter can read specific sheets
        # First, try to get sheet names using st_layers
        try:
            layers = self.conn.execute(f"SELECT * FROM st_layers('{file_path}')").fetchall()
            sheet_names = [layer[0] for layer in layers] if layers else ["Sheet1"]
        except Exception:
            # If st_layers fails, try reading without specifying a layer
            sheet_names = ["default"]

        for sheet_name in sheet_names:
            try:
                clean_sheet_name = (
                    "".join(c if c.isalnum() else "_" for c in str(sheet_name)).lower().strip("_")
                )
                if not clean_sheet_name:
                    clean_sheet_name = f"sheet_{len(sheets_data) + 1}"

                table_name = f"excel_{clean_sheet_name}_{int(time.time())}"

                if sheet_name == "default":
                    self.conn.execute(f"""
                        CREATE TABLE {table_name} AS
                        SELECT * FROM st_read('{file_path}')
                    """)
                else:
                    self.conn.execute(f"""
                        CREATE TABLE {table_name} AS
                        SELECT * FROM st_read('{file_path}', layer='{sheet_name}')
                    """)

                # Check if table has data
                row_count = self.conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                if row_count > 0:
                    sheets_data.append((clean_sheet_name, table_name))
                    logger.debug(f"Read sheet '{sheet_name}' with {row_count} rows using spatial")
                else:
                    self.drop_table(table_name)

            except Exception as e:
                logger.debug(f"Failed to read sheet '{sheet_name}' with spatial: {e}")
                continue

        return sheets_data

    def _read_excel_with_openpyxl(self, file_path: Path) -> list[tuple[str, str]]:
        """Read Excel file sheets using openpyxl, then create DuckDB tables.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of tuples containing (sheet_name, table_name)
        """
        try:
            import openpyxl

            logger.debug(f"Reading Excel file with openpyxl: {file_path}")

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets_data = []

            logger.info(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

            for sheet_name in wb.sheetnames:
                try:
                    sheet = wb[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))

                    if not rows:
                        logger.debug(f"Skipping empty sheet: {sheet_name}")
                        continue

                    # First row as headers
                    headers = []
                    for i, h in enumerate(rows[0]):
                        col_str = str(h).strip() if h is not None else ""
                        if col_str.startswith("Unnamed:") or col_str == "":
                            col_str = f"column_{i}"
                        # Replace problematic characters
                        col_str = col_str.replace('"', "'").replace("\n", " ").replace("\r", " ")
                        headers.append(col_str)

                    # Get data rows
                    data_rows = []
                    for row in rows[1:]:
                        # Skip completely empty rows
                        if all(v is None or str(v).strip() == "" for v in row):
                            continue
                        data_rows.append(row)

                    if not data_rows:
                        logger.debug(f"Skipping empty sheet after cleaning: {sheet_name}")
                        continue

                    # Clean sheet name for table name
                    clean_sheet_name = (
                        "".join(c if c.isalnum() else "_" for c in str(sheet_name))
                        .lower()
                        .strip("_")
                    )

                    if not clean_sheet_name:
                        clean_sheet_name = f"sheet_{len(sheets_data) + 1}"

                    table_name = f"excel_{clean_sheet_name}_{int(time.time())}"

                    # Create table with all columns as VARCHAR
                    columns_def = ", ".join([f'"{h}" VARCHAR' for h in headers])
                    self.conn.execute(f"CREATE TABLE {table_name} ({columns_def})")

                    # Insert data row by row
                    for row in data_rows:
                        values = []
                        for i, v in enumerate(row):
                            if i >= len(headers):
                                break
                            val = str(v) if v is not None else ""
                            # Escape single quotes for SQL
                            val = val.replace("'", "''")
                            values.append(f"'{val}'")

                        # Pad values if row is shorter than headers
                        while len(values) < len(headers):
                            values.append("''")

                        values_str = ", ".join(values)
                        self.conn.execute(f"INSERT INTO {table_name} VALUES ({values_str})")

                    sheets_data.append((clean_sheet_name, table_name))
                    logger.debug(
                        f"Read sheet '{sheet_name}' as '{clean_sheet_name}' "
                        f"with {len(data_rows)} rows"
                    )

                except Exception as e:
                    logger.warning(f"Failed to read sheet '{sheet_name}': {e!s}")
                    continue

            wb.close()
            logger.info(f"Successfully processed {len(sheets_data)} sheets from Excel file")
            return sheets_data

        except ImportError:
            logger.error("openpyxl not installed - cannot read Excel files")
            return []
        except Exception as e:
            logger.error(f"Failed to read Excel file with openpyxl: {e!s}")
            return []

    def _standardize_column_names_duckdb(self, table_name: str) -> str:
        """Standardize column names using DuckDB operations.

        Args:
            table_name: Name of source table

        Returns:
            Name of table with standardized column names
        """
        try:
            # Get current column info
            columns_info = self.get_table_info(table_name)

            if not columns_info:
                logger.warning(f"No columns found in table {table_name}")
                return table_name

            # Create column mapping for renaming
            column_mapping = []
            used_names = set()  # Track used standardized names to avoid duplicates

            for i, col_info in enumerate(columns_info):
                # DuckDB DESCRIBE returns: (col_name, col_type, nullable, key, default, extra)
                col_name = col_info[0]

                # Apply domain-specific column name mappings first
                mapped_name = self._apply_domain_specific_mappings(col_name)

                if mapped_name:
                    # Use domain-specific mapping
                    standardized_name = mapped_name
                else:
                    # Standardize column name: lowercase, replace spaces/special chars
                    # Handle Danish characters properly
                    standardized_name = str(col_name).lower()
                    # Replace Danish characters
                    standardized_name = standardized_name.replace("æ", "ae")
                    standardized_name = standardized_name.replace("ø", "oe")
                    standardized_name = standardized_name.replace("å", "aa")
                    # Replace other special characters
                    standardized_name = standardized_name.replace("é", "e")
                    standardized_name = standardized_name.replace("è", "e")
                    standardized_name = standardized_name.replace("ê", "e")
                    standardized_name = standardized_name.replace("ë", "e")
                    standardized_name = standardized_name.replace("á", "a")
                    standardized_name = standardized_name.replace("à", "a")
                    standardized_name = standardized_name.replace("â", "a")
                    standardized_name = standardized_name.replace("ä", "a")
                    standardized_name = standardized_name.replace("ó", "o")
                    standardized_name = standardized_name.replace("ò", "o")
                    standardized_name = standardized_name.replace("ô", "o")
                    standardized_name = standardized_name.replace("ö", "o")
                    # Replace spaces and special chars with underscores
                    standardized_name = "".join(
                        c if c.isalnum() else "_" for c in standardized_name
                    ).strip("_")

                    # Remove consecutive underscores
                    while "__" in standardized_name:
                        standardized_name = standardized_name.replace("__", "_")

                    # Ensure it doesn't start with a number
                    if standardized_name and standardized_name[0].isdigit():
                        standardized_name = f"col_{standardized_name}"

                    # Handle empty names or very short names
                    if not standardized_name or len(standardized_name) < 2:
                        standardized_name = f"column_{i}"

                # Ensure uniqueness
                original_standardized = standardized_name
                counter = 1
                while standardized_name in used_names:
                    standardized_name = f"{original_standardized}_{counter}"
                    counter += 1

                used_names.add(standardized_name)

                # Properly quote both original and standardized column names
                # Escape any quotes in the original column name
                escaped_col_name = str(col_name).replace('"', '""')
                column_mapping.append(f'"{escaped_col_name}" AS "{standardized_name}"')

            # Create new table with standardized column names
            result_table = f"{table_name}_clean_cols"
            columns_sql = ", ".join(column_mapping)

            self.conn.execute(f"""
                CREATE TABLE {result_table} AS
                SELECT {columns_sql}
                FROM {table_name}
            """)

            # Add backward compatibility columns for pesticide data
            self._add_backward_compatibility_columns(result_table)

            logger.debug(f"Standardized column names for table {table_name}")
            return result_table

        except Exception as e:
            logger.error(f"Failed to standardize column names: {e!s}")
            # Return original table if standardization fails
            return table_name

    def _apply_domain_specific_mappings(self, col_name: str) -> str | None:
        """Apply domain-specific column name mappings for known data types.

        Args:
            col_name: Original column name

        Returns:
            Mapped column name or None if no mapping applies
        """
        # Normalize column name for comparison (lowercase, no spaces)
        normalized_name = col_name.lower().replace(" ", "").replace("_", "")

        # Pesticide data column mappings
        pesticide_mappings = {
            "acreagesize": "area_ha",
            "companyregistrationnumber": "cvr_number",
            "code": "crop_code",
            "pesticidename": "pesticide_name",
            "pesticideregistrationnumber": "pesticide_registration_number",
            "dosagequantity": "dosage_quantity",
            "dosageunit": "dosage_unit",
            "nopesticides": "no_pesticides",
        }

        # General CVR column mappings (for any document type)
        cvr_mappings = {
            "cvr": "cvr_number",
            "cvrno": "cvr_number",
            "cvrnr": "cvr_number",
            "cvrnummer": "cvr_number",
            "virksomhedsnummer": "cvr_number",
            "companyid": "cvr_number",
            "company_id": "cvr_number",
            "firmaid": "cvr_number",
            "firma_id": "cvr_number",
        }

        # Check if this matches any pesticide column
        if normalized_name in pesticide_mappings:
            return pesticide_mappings[normalized_name]

        # Check if this matches any CVR column variation
        if normalized_name in cvr_mappings:
            return cvr_mappings[normalized_name]

        return None

    def _add_backward_compatibility_columns(self, table_name: str) -> None:
        """Add backward compatibility columns for pesticide data to maintain compatibility.

        Args:
            table_name: Name of the table to add compatibility columns to
        """
        try:
            # Get current columns to check what we're working with
            columns_info = self.get_table_info(table_name)
            column_names = [col[0] for col in columns_info]

            # Define backward compatibility mappings (new_name -> old_name)
            backward_compatibility_mappings = {
                "area_ha": "acreagesize",
                "cvr_number": "companyregistrationnumber",
                "crop_code": "code",
                "pesticide_name": "pesticidename",
                "pesticide_registration_number": "pesticideregistrationnumber",
                "dosage_quantity": "dosagequantity",
                "dosage_unit": "dosageunit",
                "no_pesticides": "nopesticides",
            }

            # Add backward compatibility columns if the new columns exist
            alter_statements = []
            for new_col, old_col in backward_compatibility_mappings.items():
                if new_col in column_names and old_col not in column_names:
                    escaped_table = table_name.replace('"', '""')
                    escaped_new = new_col.replace('"', '""')
                    escaped_old = old_col.replace('"', '""')
                    alter_statements.append(
                        f'ALTER TABLE "{escaped_table}" ADD COLUMN "{escaped_old}" '
                        f'AS ("{escaped_new}")'
                    )

            # Execute all alter statements
            for statement in alter_statements:
                self.conn.execute(statement)

            if alter_statements:
                logger.debug(
                    f"Added {len(alter_statements)} backward compatibility columns to {table_name}"
                )

        except Exception as e:
            logger.warning(f"Failed to add backward compatibility columns to {table_name}: {e!s}")

    def _standardize_data_types_duckdb(self, table_name: str) -> str:
        """Apply data type standardization using DuckDB operations.

        Args:
            table_name: Name of source table

        Returns:
            Name of table with standardized data types
        """
        try:
            # Get column information
            columns_info = self.get_table_info(table_name)

            if not columns_info:
                logger.warning(f"No columns found in table {table_name}")
                return table_name

            # Build column transformations
            column_transformations = []

            for col_info in columns_info:
                # DuckDB DESCRIBE returns: (col_name, col_type, nullable, key, default, extra)
                col_name = col_info[0]

                # Properly quote column name for DuckDB and escape quotes
                escaped_col_name = str(col_name).replace('"', '""')
                quoted_col_name = f'"{escaped_col_name}"'

                # Apply explicit type casting for known columns
                if col_name in [
                    "area_ha",
                    "block_area_ha",
                    "applied_area_ha",
                    "reported_area_ha",
                    "dosage_quantity",
                ]:
                    # Force area and dosage columns to DOUBLE with safe casting
                    transformation = (
                        f'TRY_CAST({quoted_col_name} AS DOUBLE) AS "{escaped_col_name}"'
                    )
                elif col_name in ["cvr_number", "crop_code", "pesticide_registration_number"]:
                    # Force ID columns to VARCHAR
                    transformation = f'CAST({quoted_col_name} AS VARCHAR) AS "{escaped_col_name}"'
                else:
                    # For all other columns, cast to VARCHAR to ensure consistency
                    # This avoids type inference issues and makes the data more predictable
                    transformation = f'CAST({quoted_col_name} AS VARCHAR) AS "{escaped_col_name}"'

                column_transformations.append(transformation)

            # Create new table with type standardization
            result_table = f"{table_name}_typed"
            transformations_sql = ",\n                ".join(column_transformations)

            # Simplified query without complex row filtering to avoid parsing issues
            self.conn.execute(f"""
                CREATE TABLE {result_table} AS
                SELECT
                    {transformations_sql}
                FROM {table_name}
                WHERE 1=1
            """)

            logger.debug(f"Applied data type standardization to table {table_name}")
            return result_table

        except Exception as e:
            logger.error(f"Failed to standardize data types: {e!s}")
            # Return original table if standardization fails
            return table_name
